import pandas as pd
import os
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image

def read_and_combine_data(data_directory):
    all_data = []
    file_count = 0
    for file in os.listdir(data_directory):
        if file.endswith(".xlsx"):
            file_count += 1
            file_path = os.path.join(data_directory, file)
            print(f"Reading file: {file_path}")  # Debug print to confirm file path
            try:
                data = pd.read_excel(file_path)
                print(f"Data from {file} before dropping NA and duplicates:\n{data.head()}")  # Show some data
                data.dropna(subset=['Sales'], inplace=True)
                data['Sales'] = data['Sales'].astype(float)
                data = data.drop_duplicates()
                data.columns = [x.strip().title() for x in data.columns]  # Standardize column names
                all_data.append(data)
            except Exception as e:
                print(f"Error processing file {file}: {e}")
    if file_count == 0:
        print("No Excel files found in the directory.")
    combined_data = pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()
    print(f"Combined Data Columns: {combined_data.columns}")  # Debug print to confirm column names
    return combined_data


def analyze_data(monthly_data):
    total_sales = monthly_data['Sales'].sum() if not monthly_data.empty else 0
    average_sales = monthly_data['Sales'].mean() if not monthly_data.empty else 0
    return total_sales, average_sales


def generate_plots(monthly_data, report_path):
    if 'Sales' in monthly_data.columns:
        if 'Date' in monthly_data.columns:
            plt.figure(figsize=(10, 6))
            plt.plot(monthly_data['Date'], monthly_data['Sales'], marker='o')
            plt.title('Sales Trend')
            plt.xlabel('Date')
            plt.ylabel('Sales')
            plt.grid(True)
            plt.savefig(os.path.join(report_path, 'sales_trend.png'))
            plt.close()
        else:
            print("Date column not found, skipping trend plot.")

        plt.figure(figsize=(10, 6))
        sns.histplot(monthly_data['Sales'], kde=True, color='blue')
        plt.title('Sales Distribution')
        plt.xlabel('Sales')
        plt.ylabel('Frequency')
        plt.savefig(os.path.join(report_path, 'sales_distribution.png'))
        plt.close()
    else:
        print("Sales column not found, skipping all plots.")

def generate_report_with_formatting(monthly_data, total_sales, average_sales, report_path):
    with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
        monthly_data.to_excel(writer, sheet_name='Monthly Data', index=False)
        summary_df = pd.DataFrame({
            'Metric': ['Total Sales', 'Average Daily Sales'],
            'Value': [total_sales, average_sales]
        })
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    wb = load_workbook(report_path)
    ws = wb['Monthly Data']

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    sales_values = monthly_data['Sales'].tolist()
    max_sales = max(sales_values)
    min_sales = min(sales_values)

    for idx, val in enumerate(sales_values, start=2):
        if val == max_sales:
            ws[f'B{idx}'].fill = green_fill
        elif val == min_sales:
            ws[f'B{idx}'].fill = red_fill

    img_path = os.path.join(os.path.dirname(report_path), 'sales_trend.png')
    img = Image(img_path)
    ws.add_image(img, 'D1')

    img_path = os.path.join(os.path.dirname(report_path), 'sales_distribution.png')
    img = Image(img_path)
    ws.add_image(img, 'D20')

    wb.save(report_path)

def main():
    data_directory = r"C:\Users\Admin\Documents\Spreadsheets\Sheets"
    report_path = r"C:\Users\Admin\Documents\Spreadsheets\New_Sheets\monthly_sales_report.xlsx"
    print("Reading and combining data...")
    monthly_data = read_and_combine_data(data_directory)
    print("Analyzing data...")
    total_sales, average_sales = analyze_data(monthly_data)
    print("Generating plots...")
    generate_plots(monthly_data, os.path.dirname(report_path))
    print("Generating formatted report...")
    generate_report_with_formatting(monthly_data, total_sales, average_sales, report_path)
    print("Report generated successfully.")

if __name__ == "__main__":
    main()
